#!/usr/bin/env python3
"""Parse ESG Excel and generate question-bank.ts and answer-database.ts"""

import json
import os
import sys
from pathlib import Path

import openpyxl

# Ensure project root is in path for config imports
PROJECT_ROOT = Path(__file__).resolve().parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from esggo.config import SUSTAIN_WRITE_DIR, get_full_excel  # noqa: E402

EXCEL = str(get_full_excel())
OUT = str(SUSTAIN_WRITE_DIR)


def parse_question_bank(wb) -> list:
    """Parse sheet 02 and return list of questions."""
    ws = wb['02_C版140題題庫']
    questions = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        questions.append({
            "id": str(row[0]),
            "chapter": str(row[1]),
            "question": str(row[2]),
            "what_to_fill": str(row[3])[:200] if row[3] else "",
            "why_fill": str(row[4])[:200] if row[4] else "",
            "gri_mapping": str(row[5]) if row[5] else "",
            "evidence": str(row[6]) if row[6] else "",
            "ai_help": str(row[7]) if row[7] else "",
        })
    return questions


def write_question_bank(questions: list, output_dir: str) -> None:
    """Write question-bank.ts file."""
    with open(os.path.join(output_dir, "question-bank.ts"), "w", encoding="utf-8") as f:
        f.write("export interface Question {\n")
        f.write("  readonly id: string;\n")
        f.write("  readonly chapter: string;\n")
        f.write("  readonly question: string;\n")
        f.write("  readonly whatToFill: string;\n")
        f.write("  readonly whyFill: string;\n")
        f.write("  readonly griMapping: string;\n")
        f.write("  readonly evidence: string;\n")
        f.write("  readonly aiHelp: string;\n")
        f.write("}\n\n")
        f.write("export const QUESTIONS: readonly Question[] = [\n")
        for q in questions:
            f.write("  {\n")
            f.write(f"    id: {json.dumps(q['id'])},\n")
            f.write(f"    chapter: {json.dumps(q['chapter'])},\n")
            f.write(f"    question: {json.dumps(q['question'])},\n")
            f.write(f"    whatToFill: {json.dumps(q['what_to_fill'])},\n")
            f.write(f"    whyFill: {json.dumps(q['why_fill'])},\n")
            f.write(f"    griMapping: {json.dumps(q['gri_mapping'])},\n")
            f.write(f"    evidence: {json.dumps(q['evidence'])},\n")
            f.write(f"    aiHelp: {json.dumps(q['ai_help'])},\n")
            f.write("  },\n")
        f.write("];\n\n")
        f.write("export function getQuestionsByChapter(chapter: string): Question[] {\n")
        f.write("  return QUESTIONS.filter(q => q.chapter === chapter);\n")
        f.write("}\n\n")
        f.write("export function getQuestionById(id: string): Question | undefined {\n")
        f.write("  return QUESTIONS.find(q => q.id === id);\n")
        f.write("}\n")


def parse_answers(wb) -> list:
    """Parse sheet 03 and return list of answers."""
    ws2 = wb['03_C版完整填答1400筆']
    answers = []
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, values_only=True):
        if not row[0]:
            continue
        answers.append({
            "companyId": str(row[0]),
            "companyType": str(row[1]) if row[1] else "",
            "companyName": str(row[2]) if row[2] else "",
            "questionId": str(row[3]),
            "chapter": str(row[4]) if row[4] else "",
            "question": str(row[5]) if row[5] else "",
            "answer": str(row[6])[:500] if row[6] else "",
            "dataAtoms": str(row[7]) if row[7] else "",
            "griMapping": str(row[8]) if row[8] else "",
            "evidence": str(row[9]) if row[9] else "",
            "reportDirection": str(row[10]) if row[10] else "",
            "maturity": str(row[11]) if row[11] else "",
            "gaps": str(row[12]) if row[12] else "",
        })
    return answers


def write_answer_database(answers: list, output_dir: str) -> None:
    """Write answer-database.ts file."""
    with open(os.path.join(output_dir, "answer-database.ts"), "w", encoding="utf-8") as f:
        f.write("export interface Answer {\n")
        f.write("  readonly companyId: string;\n")
        f.write("  readonly companyType: string;\n")
        f.write("  readonly companyName: string;\n")
        f.write("  readonly questionId: string;\n")
        f.write("  readonly chapter: string;\n")
        f.write("  readonly question: string;\n")
        f.write("  readonly answer: string;\n")
        f.write("  readonly dataAtoms: string;\n")
        f.write("  readonly griMapping: string;\n")
        f.write("  readonly evidence: string;\n")
        f.write("  readonly reportDirection: string;\n")
        f.write("  readonly maturity: string;\n")
        f.write("  readonly gaps: string;\n")
        f.write("}\n\n")
        f.write("export const ANSWERS: readonly Answer[] = [\n")
        for a in answers:
            f.write("  {\n")
            f.write(f"    companyId: {json.dumps(a['companyId'])},\n")
            f.write(f"    companyType: {json.dumps(a['companyType'])},\n")
            f.write(f"    companyName: {json.dumps(a['companyName'])},\n")
            f.write(f"    questionId: {json.dumps(a['questionId'])},\n")
            f.write(f"    chapter: {json.dumps(a['chapter'])},\n")
            f.write(f"    question: {json.dumps(a['question'])},\n")
            f.write(f"    answer: {json.dumps(a['answer'])},\n")
            f.write(f"    dataAtoms: {json.dumps(a['dataAtoms'])},\n")
            f.write(f"    griMapping: {json.dumps(a['griMapping'])},\n")
            f.write(f"    evidence: {json.dumps(a['evidence'])},\n")
            f.write(f"    reportDirection: {json.dumps(a['reportDirection'])},\n")
            f.write(f"    maturity: {json.dumps(a['maturity'])},\n")
            f.write(f"    gaps: {json.dumps(a['gaps'])},\n")
            f.write("  },\n")
        f.write("];\n\n")
        f.write("export function getAnswersByCompany(companyId: string): Answer[] {\n")
        f.write("  return ANSWERS.filter(a => a.companyId === companyId);\n")
        f.write("}\n\n")
        f.write("export function getAnswersByQuestion(questionId: string): Answer[] {\n")
        f.write("  return ANSWERS.filter(a => a.questionId === questionId);\n")
        f.write("}\n\n")
        f.write("export function getCompanyIds(): string[] {\n")
        f.write("  return [...new Set(ANSWERS.map(a => a.companyId))];\n")
        f.write("}\n")


def main() -> None:
    """Main entry point: parse Excel and generate TypeScript files."""
    print('Loading Excel file...')
    wb = openpyxl.load_workbook(EXCEL, data_only=True)

    # === 1. Question Bank ===
    questions = parse_question_bank(wb)
    print(f"Questions: {len(questions)}")
    write_question_bank(questions, OUT)
    print("question-bank.ts written")

    # === 2. Answer Database ===
    answers = parse_answers(wb)
    print(f"Answers: {len(answers)}")
    write_answer_database(answers, OUT)
    print("answer-database.ts written")

    wb.close()
    print("DONE")


if __name__ == '__main__':
    main()
