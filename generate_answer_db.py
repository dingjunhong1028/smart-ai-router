#!/usr/bin/env python3
"""Generate answer-database.ts from ESG report Excel file."""

import sys
from pathlib import Path

import openpyxl

# Ensure project root is in path for config imports
PROJECT_ROOT = Path(__file__).resolve().parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from esggo.config import get_answer_database_ts, get_tmp_excel  # noqa: E402

EXCEL_PATH = str(get_tmp_excel())
OUTPUT_PATH = str(get_answer_database_ts())


def escape_ts_string(s: str) -> str:
    """Escape a string for TypeScript single-quoted string literal."""
    if not s:
        return ""
    s = s.replace("\\", "\\\\")
    s = s.replace("'", "\\'")
    s = s.replace("\n", "\\n")
    s = s.replace("\r", "\\r")
    s = s.replace("\t", "\\t")
    return s


def main() -> None:
    """Main entry point: parse Excel and generate TypeScript."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # ========== 1. Parse Company Profiles ==========
    ws_profile = wb["01_10家公司Profile"]
    companies = []
    for row in range(2, ws_profile.max_row + 1):
        instance_id = ws_profile.cell(row=row, column=1).value
        if not instance_id:
            continue
        company = {
            "instanceId": str(instance_id).strip(),
            "industryType": str(ws_profile.cell(row=row, column=2).value or "").strip(),
            "companyName": str(ws_profile.cell(row=row, column=3).value or "").strip(),
            "shortName": str(ws_profile.cell(row=row, column=4).value or "").strip(),
            "employees": str(ws_profile.cell(row=row, column=6).value or "").strip(),
            "annualRevenue": str(ws_profile.cell(row=row, column=7).value or "").strip(),
            "operatingLocations": str(ws_profile.cell(row=row, column=8).value or "").strip(),
            "mainBusiness": str(ws_profile.cell(row=row, column=9).value or "").strip(),
            "electricityKwh": str(ws_profile.cell(row=row, column=11).value or "").strip(),
            "waterTons": str(ws_profile.cell(row=row, column=12).value or "").strip(),
        }
        companies.append(company)

    print(f"Parsed {len(companies)} companies")

    # ========== 2. Parse Answers ==========
    ws_answers = wb["03_C版完整填答1400筆"]
    answers = []
    instance_to_company = {}

    for row in range(2, ws_answers.max_row + 1):
        instance_id = ws_answers.cell(row=row, column=1).value
        if not instance_id:
            continue
        instance_id = str(instance_id).strip()

        company_name = str(ws_answers.cell(row=row, column=3).value or "").strip()
        question_id = str(ws_answers.cell(row=row, column=4).value or "").strip()
        chapter = str(ws_answers.cell(row=row, column=5).value or "").strip()
        answer_text = str(ws_answers.cell(row=row, column=7).value or "").strip()
        gri = str(ws_answers.cell(row=row, column=9).value or "").strip()
        direction = str(ws_answers.cell(row=row, column=11).value or "").strip()

        if instance_id not in instance_to_company:
            instance_to_company[instance_id] = company_name

        answers.append({
            "instanceId": instance_id,
            "questionId": question_id,
            "chapter": chapter,
            "answer": answer_text,
            "gri": gri,
            "direction": direction,
        })

    print(f"Parsed {len(answers)} answers")

    # ========== 3. Generate TypeScript ==========

    # Group answers by instanceId
    answers_by_company = {}
    for a in answers:
        cid = a["instanceId"]
        if cid not in answers_by_company:
            answers_by_company[cid] = []
        answers_by_company[cid].append(a)

    lines = []
    lines.append("// Auto-generated from ESG report Excel - 2026-06-26")
    lines.append("// DO NOT EDIT MANUALLY")
    lines.append(f"// Total answers: {len(answers)}")
    lines.append(f"// Total companies: {len(companies)}")
    lines.append("")
    lines.append("export interface V5Answer {")
    lines.append("  chapter: string;")
    lines.append("  questionId: string;")
    lines.append("  answer: string;")
    lines.append("  gri?: string;")
    lines.append("  direction?: string;")
    lines.append("}")
    lines.append("")
    lines.append("export interface CompanyProfile {")
    lines.append("  instanceId: string;")
    lines.append("  companyName: string;")
    lines.append("  shortName: string;")
    lines.append("  industryType: string;")
    lines.append("  employees: string;")
    lines.append("  annualRevenue: string;")
    lines.append("  operatingLocations: string;")
    lines.append("  mainBusiness: string;")
    lines.append("  electricityKwh: string;")
    lines.append("  waterTons: string;")
    lines.append("}")
    lines.append("")

    # COMPANIES array
    lines.append("export const COMPANIES: CompanyProfile[] = [")
    for c in companies:
        lines.append("  {")
        lines.append(f"    instanceId: '{escape_ts_string(c['instanceId'])}',")
        lines.append(f"    companyName: '{escape_ts_string(c['companyName'])}',")
        lines.append(f"    shortName: '{escape_ts_string(c['shortName'])}',")
        lines.append(f"    industryType: '{escape_ts_string(c['industryType'])}',")
        lines.append(f"    employees: '{escape_ts_string(c['employees'])}',")
        lines.append(f"    annualRevenue: '{escape_ts_string(c['annualRevenue'])}',")
        lines.append(f"    operatingLocations: '{escape_ts_string(c['operatingLocations'])}',")
        lines.append(f"    mainBusiness: '{escape_ts_string(c['mainBusiness'])}',")
        lines.append(f"    electricityKwh: '{escape_ts_string(c['electricityKwh'])}',")
        lines.append(f"    waterTons: '{escape_ts_string(c['waterTons'])}'")
        lines.append("  },")
    lines.append("];")
    lines.append("")

    # INSTANCE_TO_COMPANY mapping
    lines.append("const INSTANCE_TO_COMPANY: Record<string, string> = {")
    for cid, cname in sorted(instance_to_company.items()):
        lines.append(f"  '{cid}': '{escape_ts_string(cname)}',")
    lines.append("};")
    lines.append("")

    # Per-company answer arrays
    for cid in sorted(answers_by_company.keys()):
        company_answers = answers_by_company[cid]
        var_name = cid.replace("-", "_")
        lines.append(f"const ANSWERS_{var_name}: V5Answer[] = [")
        for a in company_answers:
            lines.append("  {")
            lines.append(f"    chapter: '{escape_ts_string(a['chapter'])}',")
            lines.append(f"    questionId: '{escape_ts_string(a['questionId'])}',")
            lines.append(f"    answer: '{escape_ts_string(a['answer'])}',")
            if a["gri"]:
                lines.append(f"    gri: '{escape_ts_string(a['gri'])}',")
            if a["direction"]:
                lines.append(f"    direction: '{escape_ts_string(a['direction'])}',")
            lines.append("  },")
        lines.append("];")
        lines.append("")

    # getAnswersByCompany function
    lines.append("export function getAnswersByCompany(companyId: string): V5Answer[] {")
    lines.append("  switch (companyId) {")
    for cid in sorted(answers_by_company.keys()):
        var_name = cid.replace("-", "_")
        lines.append(f"    case '{cid}': return ANSWERS_{var_name};")
    lines.append("    default: return [];")
    lines.append("  }")
    lines.append("}")
    lines.append("")

    # getCompanyName helper
    lines.append("export function getCompanyName(instanceId: string): string {")
    lines.append("  return INSTANCE_TO_COMPANY[instanceId] || \"\";")
    lines.append("}")
    lines.append("")

    output = "\n".join(lines)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write(output)

    print(f"\nGenerated: {OUTPUT_PATH}")
    print(f"File size: {len(output)} bytes")
    print(f"Total answers: {len(answers)}")
    print(f"Total companies: {len(companies)}")
    print("\nAnswers per company:")
    for cid in sorted(answers_by_company.keys()):
        print(f"  {cid}: {len(answers_by_company[cid])} answers")


if __name__ == "__main__":
    main()
