#!/usr/bin/env python3
"""ESGGO v5.0 — Quick data loader for ESG Excel."""

import json
import os
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

# Ensure project root is in path for config imports
PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from esggo.config import get_tmp_excel  # noqa: E402

EXCEL = str(get_tmp_excel())
OUTDIR = r"C:/var/www/esggo/reports"


def main() -> None:
    """Load Excel data and print summary."""
    os.makedirs(OUTDIR, exist_ok=True)
    wb = openpyxl.load_workbook(EXCEL, read_only=True)

    # Profiles
    ws = wb["01_10家公司Profile"]
    profiles = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            profiles[row[0]] = dict(
                zip(["id", "industry", "name", "short", "scale", "employees",
                      "revenue", "locations", "business", "energy", "kwh", "tons"], row)
            )

    # Answers
    ws2 = wb["03_C版完整填答1400筆"]
    answers = defaultdict(list)
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[0] and row[6]:
            answers[row[0]].append({
                "qid": row[3], "ch": row[4], "ans": row[6],
                "gri": row[8], "atoms": row[7], "dir": row[10], "mat": row[11],
            })

    wb.close()
    print(f"Loaded {len(profiles)} profiles, {len(answers)} companies")
    print(json.dumps({k: len(v) for k, v in answers.items()}))


if __name__ == '__main__':
    main()
