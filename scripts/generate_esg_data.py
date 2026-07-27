#!/usr/bin/env python3
"""Parse ESG Excel file and generate TypeScript question bank, answer database, and report templates."""

import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

# Ensure project root is in path for config imports
PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from esggo.config import SUSTAIN_WRITE_DIR, get_full_excel  # noqa: E402

EXCEL_PATH = str(get_full_excel())
OUTPUT_DIR = str(SUSTAIN_WRITE_DIR)

def sanitize_ts_string(s):
    """Escape a string for TypeScript single-quoted string literal."""
    if s is None:
        return ''
    s = str(s)
    s = s.replace('\\', '\\\\')
    s = s.replace("'", "\\'")
    s = s.replace('\n', '\\n')
    s = s.replace('\r', '\\r')
    s = s.replace('\t', '\\t')
    return s

def to_camel_case(text):
    """Convert text to camelCase."""
    # Remove special characters and split by spaces/underscores
    parts = re.split(r'[\s\-_]+', text)
    if not parts:
        return ''
    result = parts[0].lower()
    for p in parts[1:]:
        if p:
            result += p[0].upper() + p[1:].lower()
    return result

def generate_question_bank(wb):
    """Parse sheet 02 and generate question-bank.ts"""
    ws = wb['02_C版140題題庫']

    questions = []
    chapters_order = []
    chapter_questions = defaultdict(list)

    rows = list(ws.iter_rows(min_row=2, values_only=True))

    for row in rows:
        if not row[0]:
            continue

        q_id = sanitize_ts_string(row[0])
        chapter = sanitize_ts_string(row[1])
        question_text = sanitize_ts_string(row[2])
        what_to_fill = sanitize_ts_string(row[3])
        why_fill = sanitize_ts_string(row[4])
        gri_impact = sanitize_ts_string(row[5])
        evidence = sanitize_ts_string(row[6])
        ai_help = sanitize_ts_string(row[7])

        if chapter not in chapters_order:
            chapters_order.append(chapter)

        q = {
            'id': q_id,
            'chapter': chapter,
            'question': question_text,
            'whatToFill': what_to_fill,
            'whyFill': why_fill,
            'griImpact': gri_impact,
            'evidence': evidence,
            'aiHelp': ai_help,
        }
        questions.append(q)
        chapter_questions[chapter].append(q)

    # Generate TypeScript
    lines = []
    lines.append('/**')
    lines.append(' * C版專業永續報告題庫 (140題)')
    lines.append(' * 自動生成自 ESG GO C版 Excel 檔案')
    lines.append(' * 來源 sheet: 02_C版140題題庫')
    lines.append(' */')
    lines.append('')
    lines.append('export interface Question {')
    lines.append('  id: string;')
    lines.append('  chapter: string;')
    lines.append('  question: string;')
    lines.append('  whatToFill: string;')
    lines.append('  whyFill: string;')
    lines.append('  griImpact: string;')
    lines.append('  evidence: string;')
    lines.append('  aiHelp: string;')
    lines.append('}')
    lines.append('')
    lines.append('export interface ChapterQuestions {')
    lines.append('  chapter: string;')
    lines.append('  chapterCode: string;')
    lines.append('  griMapping: string;')
    lines.append('  questions: Question[];')
    lines.append('}')
    lines.append('')

    # GRI mappings per chapter
    gri_mappings = {
        'C1 組織與報告邊界': 'GRI 2-1, 2-3, 2-6',
        'C2 治理與永續管理': 'GRI 2-9, 2-10, 2-11, 2-12, 2-15, 2-16, 2-17, 2-18, 2-19, 2-20, 2-21',
        'C3 重大性與利害關係人': 'GRI 2-25, 2-26, 2-29, 3-1, 3-2, 3-3',
        'C4 經濟與誠信經營': 'GRI 2-27, 201-1, 201-2, 201-3, 201-4, 203-1, 203-2, 205-1, 205-2, 205-3, 206-1, 207-1, 207-2, 207-3, 207-4',
        'C5 能源、碳與氣候': 'GRI 302-1, 302-2, 302-3, 302-4, 302-5, 305-1, 305-2, 305-3, 305-4, 305-5, 305-6, 305-7',
        'C6 水資源與廢棄物': 'GRI 303-1, 303-2, 303-3, 303-4, 303-5, 306-1, 306-2, 306-3, 306-4, 306-5',
        'C7 生物多樣性與環境衝擊': 'GRI 304-1, 304-2, 304-3, 304-4',
        'C8 員工與人才發展': 'GRI 2-7, 2-8, 2-30, 401-1, 401-2, 401-3, 404-1, 404-2, 404-3, 405-1, 405-2',
        'C9 職安、人權與社會責任': 'GRI 403-1, 403-2, 403-3, 403-4, 403-5, 403-6, 403-7, 403-8, 403-9, 403-10, 406-1, 407-1, 408-1, 409-1, 410-1, 413-1, 413-2',
        'C10 供應鏈與產品責任': 'GRI 2-6, 204-1, 308-1, 308-2, 414-1, 414-2, 416-1, 416-2, 417-1, 417-2, 417-3, 418-1',
        'C11 Impact與投資人敘事': 'GRI 201-1, 203-1, 203-2, Impact: 財務重大性, 社會影響評估',
        'C12 查核、佐證與資料治理': 'GRI 1, GRI 2, 確信標準, ISAE 3000, AA1000',
    }

    lines.append('/** 章節定義與GRI對應 */')
    lines.append('export const CHAPTER_DEFINITIONS: { code: string; name: string; griMapping: string }[] = [')
    for ch in chapters_order:
        code = ch.split(' ')[0]
        lines.append(f"  {{ code: '{code}', name: '{ch}', griMapping: '{gri_mappings.get(ch, '')}' }},")
    lines.append('];')
    lines.append('')

    # All questions array
    lines.append('/** 完整題庫 (140題) */')
    lines.append('export const QUESTION_BANK: Question[] = [')
    for q in questions:
        lines.append('  {')
        lines.append(f"    id: '{q['id']}',")
        lines.append(f"    chapter: '{q['chapter']}',")
        lines.append(f"    question: '{q['question']}',")
        lines.append(f"    whatToFill: '{q['whatToFill']}',")
        lines.append(f"    whyFill: '{q['whyFill']}',")
        lines.append(f"    griImpact: '{q['griImpact']}',")
        lines.append(f"    evidence: '{q['evidence']}',")
        lines.append(f"    aiHelp: '{q['aiHelp']}',")
        lines.append('  },')
    lines.append('];')
    lines.append('')

    # Questions by chapter
    # Questions by chapter
    lines.append('export const QUESTIONS_BY_CHAPTER: ChapterQuestions[] = [')
    for ch in chapters_order:
        code = ch.split(' ')[0]
        gri = gri_mappings.get(ch, '')
        lines.append('  {')
        lines.append(f"    chapter: '{ch}',")
        lines.append(f"    chapterCode: '{code}',")
        lines.append(f"    griMapping: '{gri}',")
        lines.append('    questions: [')
        for q in chapter_questions[ch]:
            lines.append(f"      '{q['id']}',")
        lines.append('    ],')
        lines.append('  },')
    lines.append('];')
    lines.append('')

    # Helper functions
    lines.append('/** 根據題目ID取得題目 */')
    lines.append('export function getQuestionById(id: string): Question | undefined {')
    lines.append('  return QUESTION_BANK.find(q => q.id === id);')
    lines.append('}')
    lines.append('')
    lines.append('/** 根據章節代碼取得該章節所有題目 */')
    lines.append('export function getQuestionsByChapterCode(code: string): Question[] {')
    lines.append('  return QUESTION_BANK.filter(q => q.chapter.startsWith(code));')
    lines.append('}')
    lines.append('')
    lines.append('/** 取得所有章節代碼 */')
    lines.append('export function getAllChapterCodes(): string[] {')
    lines.append('  return CHAPTER_DEFINITIONS.map(c => c.code);')
    lines.append('}')

    return '\n'.join(lines)


def generate_answer_database(wb):
    """Parse sheet 03 and generate answer-database.ts"""
    ws = wb['03_C版完整填答1400筆']

    answers = []
    companies = {}

    rows = list(ws.iter_rows(min_row=2, values_only=True))

    for row in rows:
        if not row[0]:
            continue

        instance_id = row[0]
        company_type = row[1]
        company_name = row[2]
        question_id = row[3]
        chapter = row[4]
        question_text = row[5]
        answer_text = row[6]
        data_atom = row[7]
        gri_impact = row[8]
        evidence = row[9]
        ai_report_direction = row[10]
        data_maturity = row[11]
        data_gap = row[12]

        if company_name:
            companies[company_name] = company_type

        answers.append({
            'instanceId': instance_id,
            'companyType': company_type,
            'companyName': company_name,
            'questionId': question_id,
            'chapter': chapter,
            'question': question_text,
            'answer': answer_text,
            'dataAtom': data_atom,
            'griImpact': gri_impact,
            'evidence': evidence,
            'aiReportDirection': ai_report_direction,
            'dataMaturity': data_maturity,
            'dataGap': data_gap,
        })

    # Generate TypeScript
    lines = []
    lines.append('/**')
    lines.append(' * C版專業永續報告完整填答資料庫 (1400筆)')
    lines.append(' * 自動生成自 ESG GO C版 Excel 檔案')
    lines.append(' * 來源 sheet: 03_C版完整填答1400筆')
    lines.append(' */')
    lines.append('')
    lines.append('export interface Answer {')
    lines.append('  instanceId: string;')
    lines.append('  companyType: string;')
    lines.append('  companyName: string;')
    lines.append('  questionId: string;')
    lines.append('  chapter: string;')
    lines.append('  question: string;')
    lines.append('  answer: string;')
    lines.append('  dataAtom: string;')
    lines.append('  griImpact: string;')
    lines.append('  evidence: string;')
    lines.append('  aiReportDirection: string;')
    lines.append('  dataMaturity: string;')
    lines.append('  dataGap: string;')
    lines.append('}')
    lines.append('')
    lines.append('export interface CompanyProfile {')
    lines.append('  name: string;')
    lines.append('  type: string;')
    lines.append('}')
    lines.append('')
    lines.append('export interface CompanyAnswers {')
    lines.append('  company: CompanyProfile;')
    lines.append('  answers: Answer[];')
    lines.append('  answersByChapter: Record<string, Answer[]>;')
    lines.append('}')
    lines.append('')

    # Company list
    lines.append('/** 10家模擬公司列表 */')
    lines.append('export const COMPANIES: CompanyProfile[] = [')
    for name, ctype in sorted(companies.items()):
        lines.append(f"  {{ name: '{sanitize_ts_string(name)}', type: '{sanitize_ts_string(ctype)}' }},")
    lines.append('];')
    lines.append('')

    # All answers
    lines.append('/** 完整填答資料庫 (1400筆) */')
    lines.append('export const ANSWER_DATABASE: Answer[] = [')
    for a in answers:
        lines.append('  {')
        lines.append(f"    instanceId: '{sanitize_ts_string(a['instanceId'])}',")
        lines.append(f"    companyType: '{sanitize_ts_string(a['companyType'])}',")
        lines.append(f"    companyName: '{sanitize_ts_string(a['companyName'])}',")
        lines.append(f"    questionId: '{sanitize_ts_string(a['questionId'])}',")
        lines.append(f"    chapter: '{sanitize_ts_string(a['chapter'])}',")
        lines.append(f"    question: '{sanitize_ts_string(a['question'])}',")
        lines.append(f"    answer: '{sanitize_ts_string(a['answer'])}',")
        lines.append(f"    dataAtom: '{sanitize_ts_string(a['dataAtom'])}',")
        lines.append(f"    griImpact: '{sanitize_ts_string(a['griImpact'])}',")
        lines.append(f"    evidence: '{sanitize_ts_string(a['evidence'])}',")
        lines.append(f"    aiReportDirection: '{sanitize_ts_string(a['aiReportDirection'])}',")
        lines.append(f"    dataMaturity: '{sanitize_ts_string(a['dataMaturity'])}',")
        lines.append(f"    dataGap: '{sanitize_ts_string(a['dataGap'])}',")
        lines.append('  },')
    lines.append('];')
    lines.append('')

    # Helper functions
    lines.append('/** 根據公司名稱取得所有填答 */')
    lines.append('export function getAnswersByCompany(companyName: string): Answer[] {')
    lines.append('  return ANSWER_DATABASE.filter(a => a.companyName === companyName);')
    lines.append('}')
    lines.append('')
    lines.append('/** 根據題目ID取得所有公司的填答 */')
    lines.append('export function getAnswersByQuestion(questionId: string): Answer[] {')
    lines.append('  return ANSWER_DATABASE.filter(a => a.questionId === questionId);')
    lines.append('}')
    lines.append('')
    lines.append('/** 取得特定公司特定題目的填答 */')
    lines.append('export function getAnswer(companyName: string, questionId: string): Answer | undefined {')
    lines.append('  return ANSWER_DATABASE.find(a => a.companyName === companyName && a.questionId === questionId);')
    lines.append('}')
    lines.append('')
    lines.append('/** 取得特定公司按章節分組的所有填答 */')
    lines.append('export function getCompanyAnswersByChapter(companyName: string): Record<string, Answer[]> {')
    lines.append('  const answers = getAnswersByCompany(companyName);')
    lines.append('  const byChapter: Record<string, Answer[]> = {};')
    lines.append('  for (const a of answers) {')
    lines.append('    if (!byChapter[a.chapter]) byChapter[a.chapter] = [];')
    lines.append('    byChapter[a.chapter].push(a);')
    lines.append('  }')
    lines.append('  return byChapter;')
    lines.append('}')
    lines.append('')
    lines.append('/** 取得完整公司填答資料 (含按章節分組) */')
    lines.append('export function getFullCompanyAnswers(companyName: string): CompanyAnswers {')
    lines.append('  const company = COMPANIES.find(c => c.name === companyName);')
    lines.append('  if (!company) throw new Error(`Company not found: ${companyName}`);')
    lines.append('  const answers = getAnswersByCompany(companyName);')
    lines.append('  const answersByChapter = getCompanyAnswersByChapter(companyName);')
    lines.append('  return { company, answers, answersByChapter };')
    lines.append('}')
    lines.append('')
    lines.append('/** 取得資料成熟度統計 */')
    lines.append('export function getDataMaturityStats(companyName: string): Record<string, number> {')
    lines.append('  const answers = getAnswersByCompany(companyName);')
    lines.append('  const stats: Record<string, number> = {};')
    lines.append('  for (const a of answers) {')
    lines.append('    stats[a.dataMaturity] = (stats[a.dataMaturity] || 0) + 1;')
    lines.append('  }')
    lines.append('  return stats;')
    lines.append('}')
    lines.append('')
    lines.append('/** 取得資料缺口列表 */')
    lines.append('export function getDataGaps(companyName: string): Answer[] {')
    lines.append('  return getAnswersByCompany(companyName).filter(a => a.dataGap && a.dataGap.length > 0);')
    lines.append('}')

    return '\n'.join(lines)


def generate_report_templates():
    """Generate report-templates.ts with chapter definitions and templates"""

    lines = []
    lines.append('/**')
    lines.append(' * C版專業永續報告範本與章節定義')
    lines.append(' * 自動生成自 ESG GO C版 Excel 資料庫')
    lines.append(' * 包含12章節定義、GRI對應、報告段落範本與組合函式')
    lines.append(' */')
    lines.append('')
    lines.append("import { Answer, getFullCompanyAnswers } from './answer-database';")
    lines.append('')
    lines.append('export interface ChapterDefinition {')
    lines.append('  code: string;')
    lines.append('  name: string;')
    lines.append('  title: string;')
    lines.append('  griMapping: string[];')
    lines.append('  description: string;')
    lines.append('}')
    lines.append('')
    lines.append('export interface ReportSectionTemplate {')
    lines.append('  chapterCode: string;')
    lines.append('  sectionTitle: string;')
    lines.append('  template: string;')
    lines.append('  requiredFields: string[];')
    lines.append('}')
    lines.append('')
    lines.append('export interface AssembledReport {')
    lines.append('  companyName: string;')
    lines.append('  companyType: string;')
    lines.append('  title: string;')
    lines.append('  generatedAt: string;')
    lines.append('  sections: ReportSection[];')
    lines.append('  dataMaturitySummary: Record<string, number>;')
    lines.append('  dataGaps: string[];')
    lines.append('}')
    lines.append('')
    lines.append('export interface ReportSection {')
    lines.append('  chapterCode: string;')
    lines.append('  sectionTitle: string;')
    lines.append('  content: string;')
    lines.append('  griReferences: string[];')
    lines.append('  evidenceRequired: string[];')
    lines.append('}')
    lines.append('')

    # Chapter definitions
    lines.append('/**')
    lines.append(' * C版專業永續報告 12章節定義')
    lines.append(' * 對應 GRI Standards 與 IFRS S1/S2 要求')
    lines.append(' */')
    lines.append('export const CHAPTER_DEFINITIONS: ChapterDefinition[] = [')
    lines.append('  {')
    lines.append("    code: 'C1',")
    lines.append("    name: '組織與報告邊界',")
    lines.append("    title: '第一章：組織與報告邊界',")
    lines.append("    griMapping: ['GRI 2-1', 'GRI 2-3', 'GRI 2-6'],")
    lines.append("    description: '公司組織識別、法律結構、報告邊界與聯絡窗口之完整揭露',")
    lines.append('  },')
    lines.append('  {')
    lines.append("    code: 'C2',")
    lines.append("    name: '治理與永續管理',")
    lines.append("    title: '第二章：治理與永續管理',")
    lines.append("    griMapping: ['GRI 2-9', 'GRI 2-10', 'GRI 2-11', 'GRI 2-12', 'GRI 2-15', 'GRI 2-16', 'GRI 2-17', 'GRI 2-18', 'GRI 2-19', 'GRI 2-20', 'GRI 2-21'],")
    lines.append("    description: '治理架構、永續策略、風險管理、薪酬與績效連結',")
    lines.append('  },')
    lines.append('  {')
    lines.append("    code: 'C3',")
    lines.append("    name: '重大性與利害關係人',")
    lines.append("    title: '第三章：重大性與利害關係人',")
    lines.append("    griMapping: ['GRI 2-25', 'GRI 2-26', 'GRI 2-29', 'GRI 3-1', 'GRI 3-2', 'GRI 3-3'],")
    lines.append("    description: '利害關係人辨識、重大性分析流程與結果、ESG議題優先順序',")
    lines.append('  },')
    lines.append('  {')
    lines.append("    code: 'C4',")
    lines.append("    name: '經濟與誠信經營',")
    lines.append("    title: '第四章：經濟與誠信經營',")
    lines.append("    griMapping: ['GRI 2-27', 'GRI 201-1', 'GRI 201-2', 'GRI 201-3', 'GRI 201-4', 'GRI 203-1', 'GRI 203-2', 'GRI 205-1', 'GRI 205-2', 'GRI 205-3', 'GRI 206-1', 'GRI 207-1', 'GRI 207-2', 'GRI 207-3', 'GRI 207-4'],")
    lines.append("    description: '經濟績效、市場占有率、間接經濟衝擊、反貪腐與公平競爭',")
    lines.append('  },')
    lines.append('  {')
    lines.append("    code: 'C5',")
    lines.append("    name: '能源、碳與氣候',")
    lines.append("    title: '第五章：能源、碳與氣候',")
    lines.append("    griMapping: ['GRI 302-1', 'GRI 302-2', 'GRI 302-3', 'GRI 302-4', 'GRI 302-5', 'GRI 305-1', 'GRI 305-2', 'GRI 305-3', 'GRI 305-4', 'GRI 305-5', 'GRI 305-6', 'GRI 305-7'],")
    lines.append("    description: '能源使用效率、溫室氣體排放、氣候風險與減碳目標',")
    lines.append('  },')
    lines.append('  {')
    lines.append("    code: 'C6',")
    lines.append("    name: '水資源與廢棄物',")
    lines.append("    title: '第六章：水資源與廢棄物',")
    lines.append("    griMapping: ['GRI 303-1', 'GRI 303-2', 'GRI 303-3', 'GRI 303-4', 'GRI 303-5', 'GRI 306-1', 'GRI 306-2', 'GRI 306-3', 'GRI 306-4', 'GRI 306-5'],")
    lines.append("    description: '水資源管理、用水效率、廢棄物產生與處理、循環經濟作為',")
    lines.append('  },')
    lines.append('  {')
    lines.append("    code: 'C7',")
    lines.append("    name: '生物多樣性與環境衝擊',")
    lines.append("    title: '第七章：生物多樣性與環境衝擊',")
    lines.append("    griMapping: ['GRI 304-1', 'GRI 304-2', 'GRI 304-3', 'GRI 304-4'],")
    lines.append("    description: '生物多樣性保護、生態系統影響評估、自然相關財務揭露',")
    lines.append('  },')
    lines.append('  {')
    lines.append("    code: 'C8',")
    lines.append("    name: '員工與人才發展',")
    lines.append("    title: '第八章：員工與人才發展',")
    lines.append("    griMapping: ['GRI 2-7', 'GRI 2-8', 'GRI 2-30', 'GRI 401-1', 'GRI 401-2', 'GRI 401-3', 'GRI 404-1', 'GRI 404-2', 'GRI 404-3', 'GRI 405-1', 'GRI 405-2'],")
    lines.append("    description: '人力結構、薪酬福利、人才招募與留任、訓練發展',")
    lines.append('  },')
    lines.append('  {')
    lines.append("    code: 'C9',")
    lines.append("    name: '職安、人權與社會責任',")
    lines.append("    title: '第九章：職安、人權與社會責任',")
    lines.append("    griMapping: ['GRI 403-1', 'GRI 403-2', 'GRI 403-3', 'GRI 403-4', 'GRI 403-5', 'GRI 403-6', 'GRI 403-7', 'GRI 403-8', 'GRI 403-9', 'GRI 403-10', 'GRI 406-1', 'GRI 407-1', 'GRI 408-1', 'GRI 409-1', 'GRI 410-1', 'GRI 413-1', 'GRI 413-2'],")
    lines.append("    description: '職業安全衛生、人權盡職調查、勞動實踐、社區參與',")
    lines.append('  },')
    lines.append('  {')
    lines.append("    code: 'C10',")
    lines.append("    name: '供應鏈與產品責任',")
    lines.append("    title: '第十章：供應鏈與產品責任',")
    lines.append("    griMapping: ['GRI 2-6', 'GRI 204-1', 'GRI 308-1', 'GRI 308-2', 'GRI 414-1', 'GRI 414-2', 'GRI 416-1', 'GRI 416-2', 'GRI 417-1', 'GRI 417-2', 'GRI 417-3', 'GRI 418-1'],")
    lines.append("    description: '供應鏈管理、產品安全與品質、客戶隱私與行銷溝通',")
    lines.append('  },')
    lines.append('  {')
    lines.append("    code: 'C11',")
    lines.append("    name: 'Impact與投資人敘事',")
    lines.append("    title: '第十一章：Impact與投資人敘事',")
    lines.append("    griMapping: ['GRI 201-1', 'GRI 203-1', 'GRI 203-2', 'Impact: 財務重大性', 'Impact: 社會影響評估'],")
    lines.append("    description: '影響力評估、投資人關注議題、ESG績效與財務連結',")
    lines.append('  },')
    lines.append('  {')
    lines.append("    code: 'C12',")
    lines.append("    name: '查核、佐證與資料治理',")
    lines.append("    title: '第十二章：查核、佐證與資料治理',")
    lines.append("    griMapping: ['GRI 1', 'GRI 2', 'ISAE 3000', 'AA1000'],")
    lines.append("    description: '資料治理架構、確信範圍與結果、佐證文件管理與外部查證',")
    lines.append('  },')
    lines.append('];')
    lines.append('')

    # Report section templates
    lines.append('/**')
    lines.append(' * 報告段落範本 (使用 {{placeholder}} 語法)')
    lines.append(' * 每章節對應一個範本，實際報告由 AI 根據填答內容填充')
    lines.append(' */')
    lines.append('export const REPORT_SECTION_TEMPLATES: ReportSectionTemplate[] = [')
    lines.append('  {')
    lines.append("    chapterCode: 'C1',")
    lines.append("    sectionTitle: '組織與報告邊界',")
    lines.append("    template: `## 組織與報告邊界\\n\\n本章節依據 GRI 2-1、GRI 2-3 之要求，完整揭露公司組織識別、法律結構及報告邊界。\\n\\n### 公司識別\\n{{companyName}}（{{companyType}}）設立於{{foundingYear}}，總部位於{{headquartersAddress}}，統編{{registrationNumber}}。主要營運據點包括{{operatingLocations}}。\\n\\n### 報告邊界\\n本次永續報告涵蓋{{reportingPeriod}}，發布日期為{{publishDate}}，報告週期為{{reportingCycle}}。\\n\\n### 聯絡窗口\\n報告聯絡人為{{contactPerson}}，所屬部門{{contactDepartment}}，聯絡方式：{{contactEmail}}、{{contactPhone}}。`,")
    lines.append("    requiredFields: ['C1-01', 'C1-02', 'C1-03', 'C1-04'],")
    lines.append('  },')
    lines.append('  {')
    lines.append("    chapterCode: 'C2',")
    lines.append("    sectionTitle: '治理與永續管理',")
    lines.append("    template: `## 治理與永續管理\\n\\n本章節依據 GRI 2-9 至 GRI 2-21 之要求，揭露公司治理架構、永續管理策略與績效連結機制。\\n\\n### 治理架構\\n{{governanceStructure}}\\n\\n### 永續管理策略\\n{{sustainabilityStrategy}}\\n\\n### 風險管理\\n{{riskManagement}}\\n\\n### 薪酬與績效連結\\n{{compensationLinkage}}`,")
    lines.append("    requiredFields: ['C2-01', 'C2-02', 'C2-03', 'C2-04', 'C2-05'],")
    lines.append('  },')
    lines.append('  {')
    lines.append("    chapterCode: 'C3',")
    lines.append("    sectionTitle: '重大性與利害關係人',")
    lines.append("    template: `## 重大性與利害關係人\\n\\n本章節依據 GRI 2-25、GRI 2-26、GRI 2-29、GRI 3-1 至 GRI 3-3 之要求，說明利害關係人辨識與重大性分析。\\n\\n### 利害關係人辨識\\n{{stakeholderIdentification}}\\n\\n### 重大性分析\\n{{materialityAnalysis}}\\n\\n### 重大議題\\n{{materialTopics}}`,")
    lines.append("    requiredFields: ['C3-01', 'C3-02', 'C3-03', 'C3-04', 'C3-05', 'C3-06'],")
    lines.append('  },')
    lines.append('  {')
    lines.append("    chapterCode: 'C4',")
    lines.append("    sectionTitle: '經濟與誠信經營',")
    lines.append("    template: `## 經濟與誠信經營\\n\\n本章節依據 GRI 201、GRI 203、GRI 205、GRI 206、GRI 207 之要求，揭露公司經濟績效與誠信經營作為。\\n\\n### 經濟績效\\n{{economicPerformance}}\\n\\n### 市場占有率\\n{{marketPresence}}\\n\\n### 間接經濟衝擊\\n{{indirectEconomicImpacts}}\\n\\n### 反貪腐與公平競爭\\n{{antiCorruption}}`,")
    lines.append("    requiredFields: ['C4-01', 'C4-02', 'C4-03', 'C4-04', 'C4-05'],")
    lines.append('  },')
    lines.append('  {')
    lines.append("    chapterCode: 'C5',")
    lines.append("    sectionTitle: '能源、碳與氣候',")
    lines.append("    template: `## 能源、碳與氣候\\n\\n本章節依據 GRI 302、GRI 305 之要求，揭露公司能源使用、溫室氣體排放與氣候風險管理。\\n\\n### 能源使用\\n{{energyConsumption}}\\n\\n### 溫室氣體排放\\n{{ghgEmissions}}\\n\\n### 減碳目標\\n{{carbonReductionTargets}}\\n\\n### 氣候風險\\n{{climateRisks}}`,")
    lines.append("    requiredFields: ['C5-01', 'C5-02', 'C5-03', 'C5-04', 'C5-05', 'C5-06', 'C5-07', 'C5-08'],")
    lines.append('  },')
    lines.append('  {')
    lines.append("    chapterCode: 'C6',")
    lines.append("    sectionTitle: '水資源與廢棄物',")
    lines.append("    template: `## 水資源與廢棄物\\n\\n本章節依據 GRI 303、GRI 306 之要求，揭露公司水資源管理與廢棄物處理。\\n\\n### 水資源管理\\n{{waterManagement}}\\n\\n### 用水效率\\n{{waterEfficiency}}\\n\\n### 廢棄物管理\\n{{wasteManagement}}\\n\\n### 循環經濟\\n{{circularEconomy}}`,")
    lines.append("    requiredFields: ['C6-01', 'C6-02', 'C6-03', 'C6-04', 'C6-05'],")
    lines.append('  },')
    lines.append('  {')
    lines.append("    chapterCode: 'C7',")
    lines.append("    sectionTitle: '生物多樣性與環境衝擊',")
    lines.append("    template: `## 生物多樣性與環境衝擊\\n\\n本章節依據 GRI 304 之要求，揭露公司對生物多樣性之影響與保護作為。\\n\\n### 生物多樣性評估\\n{{biodiversityAssessment}}\\n\\n### 保護措施\\n{{conservationMeasures}}\\n\\n### 環境影響評估\\n{{environmentalImpact}}`,")
    lines.append("    requiredFields: ['C7-01', 'C7-02', 'C7-03', 'C7-04'],")
    lines.append('  },')
    lines.append('  {')
    lines.append("    chapterCode: 'C8',")
    lines.append("    sectionTitle: '員工與人才發展',")
    lines.append("    template: `## 員工與人才發展\\n\\n本章節依據 GRI 2-7、GRI 401、GRI 404、GRI 405 之要求，揭露公司人力結構與人才發展策略。\\n\\n### 人力結構\\n{{workforceStructure}}\\n\\n### 薪酬福利\\n{{compensationBenefits}}\\n\\n### 人才招募與留任\\n{{talentRecruitment}}\\n\\n### 訓練發展\\n{{trainingDevelopment}}`,")
    lines.append("    requiredFields: ['C8-01', 'C8-02', 'C8-03', 'C8-04', 'C8-05', 'C8-06', 'C8-07'],")
    lines.append('  },')
    lines.append('  {')
    lines.append("    chapterCode: 'C9',")
    lines.append("    sectionTitle: '職安、人權與社會責任',")
    lines.append("    template: `## 職安、人權與社會責任\\n\\n本章節依據 GRI 403、GRI 406 至 GRI 413 之要求，揭露公司職業安全、人權與社會責任作為。\\n\\n### 職業安全衛生\\n{{occupationalSafety}}\\n\\n### 人權盡職調查\\n{{humanRightsDueDiligence}}\\n\\n### 勞動實踐\\n{{laborPractices}}\\n\\n### 社區參與\\n{{communityEngagement}}`,")
    lines.append("    requiredFields: ['C9-01', 'C9-02', 'C9-03', 'C9-04', 'C9-05', 'C9-06'],")
    lines.append('  },')
    lines.append('  {')
    lines.append("    chapterCode: 'C10',")
    lines.append("    sectionTitle: '供應鏈與產品責任',")
    lines.append("    template: `## 供應鏈與產品責任\\n\\n本章節依據 GRI 204、GRI 308、GRI 414、GRI 416 至 GRI 418 之要求，揭露公司供應鏈管理與產品責任。\\n\\n### 供應鏈管理\\n{{supplyChainManagement}}\\n\\n### 供應商永續評估\\n{{supplierSustainability}}\\n\\n### 產品安全與品質\\n{{productSafety}}\\n\\n### 客戶隱私\\n{{customerPrivacy}}`,")
    lines.append("    requiredFields: ['C10-01', 'C10-02', 'C10-03', 'C10-04', 'C10-05'],")
    lines.append('  },')
    lines.append('  {')
    lines.append("    chapterCode: 'C11',")
    lines.append("    sectionTitle: 'Impact與投資人敘事',")
    lines.append("    template: `## Impact與投資人敘事\\n\\n本章節整合 GRI 201、GRI 203 與 Impact 評估框架，說明公司ESG績效對投資人之意義。\\n\\n### 影響力評估\\n{{impactAssessment}}\\n\\n### 投資人關注議題\\n{{investorTopics}}\\n\\n### ESG績效與財務連結\\n{{esgFinancialLinkage}}`,")
    lines.append("    requiredFields: ['C11-01', 'C11-02', 'C11-03'],")
    lines.append('  },')
    lines.append('  {')
    lines.append("    chapterCode: 'C12',")
    lines.append("    sectionTitle: '查核、佐證與資料治理',")
    lines.append("    template: `## 查核、佐證與資料治理\\n\\n本章節依據 GRI 1、GRI 2 與 ISAE 3000 之要求，說明公司資料治理與外部查證機制。\\n\\n### 資料治理架構\\n{{dataGovernance}}\\n\\n### 確信範圍與結果\\n{{assuranceScope}}\\n\\n### 佐證文件管理\\n{{evidenceManagement}}\\n\\n### 外部查證\\n{{externalVerification}}`,")
    lines.append("    requiredFields: ['C12-01', 'C12-02', 'C12-03', 'C12-04'],")
    lines.append('  },')
    lines.append('];')
    lines.append('')

    # Assemble report function
    lines.append('/**')
    lines.append(' * 組合完整永續報告')
    lines.append(' * @param companyName - 公司名稱')
    lines.append(' * @returns 組合後的報告物件')
    lines.append(' */')
    lines.append('export function assembleReport(companyName: string): AssembledReport {')
    lines.append('  const { company, answers, answersByChapter } = getFullCompanyAnswers(companyName);')
    lines.append('')
    lines.append('  const sections: ReportSection[] = [];')
    lines.append('  const dataGaps: string[] = [];')
    lines.append('')
    lines.append('  for (const template of REPORT_SECTION_TEMPLATES) {')
    lines.append('    const chapterAnswers = answersByChapter[template.chapterCode] || [];')
    lines.append('    const answerMap = new Map<string, Answer>();')
    lines.append('    for (const a of chapterAnswers) {')
    lines.append('      answerMap.set(a.questionId, a);')
    lines.append('    }')
    lines.append('')
    lines.append('    // Build content from answers')
    lines.append('    let content = template.template;')
    lines.append('    const griRefs: string[] = [];')
    lines.append('    const evidenceReq: string[] = [];')
    lines.append('')
    lines.append('    for (const answer of chapterAnswers) {')
    lines.append('      if (answer.griImpact) {')
    lines.append('        for (const gri of answer.griImpact.split(/[,、]/)) {')
    lines.append('          const trimmed = gri.trim();')
    lines.append('          if (trimmed && !griRefs.includes(trimmed)) griRefs.push(trimmed);')
    lines.append('        }')
    lines.append('      }')
    lines.append('      if (answer.evidence) {')
    lines.append('        for (const ev of answer.evidence.split(/[;；]/)) {')
    lines.append('          const trimmed = ev.trim();')
    lines.append('          if (trimmed && !evidenceReq.includes(trimmed)) evidenceReq.push(trimmed);')
    lines.append('        }')
    lines.append('      }')
    lines.append('      if (answer.dataGap) {')
    lines.append('        dataGaps.push(`[${answer.questionId}] ${answer.dataGap}`);')
    lines.append('      }')
    lines.append('    }')
    lines.append('')
    lines.append('    // Replace {{companyName}} and {{companyType}}')
    lines.append("    content = content.replace(/\\{\\{companyName\\}\\}/g, company.name);")
    lines.append("    content = content.replace(/\\{\\{companyType\\}\\}/g, company.type);")
    lines.append('')
    lines.append('    // Replace {{answer:questionId}} with actual answer')
    lines.append("    content = content.replace(/\\{\\{answer:([^}]+)\\}\\}/g, (_, qId: string) => {")
    lines.append('      const a = answerMap.get(qId);')
    lines.append('      return a ? a.answer : `[待填充: ${qId}]`;')
    lines.append('    });')
    lines.append('')
    lines.append('    // Replace {{field:questionId}} with first line of answer')
    lines.append("    content = content.replace(/\\{\\{field:([^}]+)\\}\\}/g, (_, qId: string) => {")
    lines.append('      const a = answerMap.get(qId);')
    lines.append('      if (!a) return `[待填充: ${qId}]`;')
    lines.append('      return a.answer.split(/[。\n]/)[0] || a.answer.substring(0, 100);')
    lines.append('    });')
    lines.append('')
    lines.append('    sections.push({')
    lines.append('      chapterCode: template.chapterCode,')
    lines.append('      sectionTitle: template.sectionTitle,')
    lines.append('      content,')
    lines.append('      griReferences: griRefs,')
    lines.append('      evidenceRequired: evidenceReq,')
    lines.append('    });')
    lines.append('  }')
    lines.append('')
    lines.append('  // Build data maturity summary')
    lines.append('  const maturitySummary: Record<string, number> = {};')
    lines.append('  for (const a of answers) {')
    lines.append('    maturitySummary[a.dataMaturity] = (maturitySummary[a.dataMaturity] || 0) + 1;')
    lines.append('  }')
    lines.append('')
    lines.append('  return {')
    lines.append('    companyName: company.name,')
    lines.append('    companyType: company.type,')
    lines.append('    title: `${company.name} 永續報告 (C版專業揭露)`,')
    lines.append('    generatedAt: new Date().toISOString(),')
    lines.append('    sections,')
    lines.append('    dataMaturitySummary: maturitySummary,')
    lines.append('    dataGaps,')
    lines.append('  };')
    lines.append('}')
    lines.append('')
    lines.append('/**')
    lines.append(' * 將組合後的報告轉為 Markdown 字串')
    lines.append(' */')
    lines.append('export function reportToMarkdown(report: AssembledReport): string {')
    lines.append('  const parts: string[] = [];')
    lines.append('')
    lines.append('  parts.push(`# ${report.title}`);')
    lines.append('  parts.push(`\\n> 報告類型：${report.companyType}`);')
    lines.append("  parts.push(`> 產生時間：${report.generatedAt}\\n`);")
    lines.append('')
    lines.append('  // Table of contents')
    lines.append('  parts.push(`## 目錄\\n`);')
    lines.append('  for (const section of report.sections) {')
    lines.append('    parts.push(`- ${section.chapterCode} ${section.sectionTitle}`);')
    lines.append('  }')
    lines.append("  parts.push('\\n---\\n');")
    lines.append('')
    lines.append('  // Sections')
    lines.append('  for (const section of report.sections) {')
    lines.append('    parts.push(section.content);')
    lines.append("    parts.push('\\n\\n### GRI 參考\\n');")
    lines.append("    parts.push(section.griReferences.map(g => `  - ${g}`).join('\\n'));")
    lines.append("    parts.push('\\n\\n### 需要佐證\\n');")
    lines.append("    parts.push(section.evidenceRequired.map(e => `  - ${e}`).join('\\n'));")
    lines.append("    parts.push('\\n---\\n');")
    lines.append('  }')
    lines.append('')
    lines.append('  // Data maturity summary')
    lines.append('  parts.push(`## 資料成熟度統計\\n`);')
    lines.append('  for (const [level, count] of Object.entries(report.dataMaturitySummary)) {')
    lines.append('    parts.push(`- ${level}: ${count} 題`);')
    lines.append('  }')
    lines.append("  parts.push('\\n');")
    lines.append('')
    lines.append('  // Data gaps')
    lines.append('  if (report.dataGaps.length > 0) {')
    lines.append("    parts.push('## 資料缺口\\n');")
    lines.append('    for (const gap of report.dataGaps) {')
    lines.append('      parts.push(`- ${gap}`);')
    lines.append('    }')
    lines.append('  }')
    lines.append('')
    lines.append("  return parts.join('\\n');")
    lines.append('}')
    lines.append('')
    lines.append('/**')
    lines.append(' * 匯出報告為 JSON 格式')
    lines.append(' */')
    lines.append('export function reportToJSON(report: AssembledReport): string {')
    lines.append('  return JSON.stringify(report, null, 2);')
    lines.append('}')
    lines.append('')
    lines.append('/** 取得所有章節定義 */')
    lines.append('export function getChapterDefinitions(): ChapterDefinition[] {')
    lines.append('  return CHAPTER_DEFINITIONS;')
    lines.append('}')
    lines.append('')
    lines.append('/** 取得特定章節的報告範本 */')
    lines.append('export function getChapterTemplate(chapterCode: string): ReportSectionTemplate | undefined {')
    lines.append('  return REPORT_SECTION_TEMPLATES.find(t => t.chapterCode === chapterCode);')
    lines.append('}')

    return '\n'.join(lines)


def main():
    print('Loading Excel file...')
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    print(f'Sheet names: {wb.sheetnames}')

    # Generate question bank
    print('Generating question-bank.ts...')
    qb_content = generate_question_bank(wb)
    qb_path = f'{OUTPUT_DIR}\\question-bank.ts'
    with open(qb_path, 'w', encoding='utf-8') as f:
        f.write(qb_content)
    print(f'  Written: {qb_path} ({len(qb_content)} chars)')

    # Generate answer database
    print('Generating answer-database.ts...')
    ad_content = generate_answer_database(wb)
    ad_path = f'{OUTPUT_DIR}\\answer-database.ts'
    with open(ad_path, 'w', encoding='utf-8') as f:
        f.write(ad_content)
    print(f'  Written: {ad_path} ({len(ad_content)} chars)')

    # Generate report templates
    print('Generating report-templates.ts...')
    rt_content = generate_report_templates()
    rt_path = f'{OUTPUT_DIR}\\report-templates.ts'
    with open(rt_path, 'w', encoding='utf-8') as f:
        f.write(rt_content)
    print(f'  Written: {rt_path} ({len(rt_content)} chars)')

    wb.close()
    print('\nDone! All files generated successfully.')


if __name__ == '__main__':
    main()
